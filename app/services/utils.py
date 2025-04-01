#Bloc 1 — Fonctions Selenium & HTML

from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import pycountry
from pytz import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from io import StringIO
import os

def get_element_html(driver, xpath):
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
    return element.get_attribute('outerHTML')

def find_div_with_viewbox(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    divs = soup.find_all(lambda tag: tag.name == 'div' and 'viewBoxe' in tag.get('class', []) and not any(
        input_tag.get('id', '').lower().find('mouillage') != -1 for input_tag in tag.find_all('input')))
    return divs

def extract_info_from_divs(divs):
    data = []
    for div in divs:
        try:
            lloyd = div.find('input', id=lambda x: x and 'hidLloyd' in x).get('value')
            nom_navire = div.find('a', id=lambda x: x and "lnkNomNavire" in x).get_text()
            type_navire = div.find('span', id=lambda x: x and 'Type' in x).get_text()
            operation = div.find('span', id=lambda x: x and 'Qualifiant' in x).get_text()
            prov_or_dest = div.find('span', id=lambda x: x and 'PortProvDest' in x).get_text()
            date_et_heure = div.find('input', id=lambda x: x and 'hidEtd' in x).get('value')
            mouvement = div.find('input', id=lambda x: x and 'hidSens' in x).get('value')
            data.append([lloyd, nom_navire, type_navire, operation, prov_or_dest, date_et_heure, mouvement])
        except:
            continue
    return pd.DataFrame(data, columns=['Lloyd', 'Nom Navire', 'Type Navire', 'Opération', 'Provenance ou Destination', 'Date et Heure', 'Mouvement'])




#Bloc 2 — Fonctions de traitement des données

def supprimer_D(df):
    return df[df['Mouvement'] != 'D']

def convertion(df):
    df = df.copy()
    df['Arrivée'] = pd.to_datetime(df['Arrivée'], dayfirst=True, errors='raise')
    df['Arrivée'] = df['Arrivée'].apply(lambda x: timezone('Europe/Paris').localize(x).astimezone(timezone('UTC')))
    df['Arrivée'] = df['Arrivée'].dt.tz_localize(None)
    return df

def transformer_date(df):
    df['Arrivée'] = pd.to_datetime(df['Arrivée'], dayfirst=True, errors='raise')
    df['Arrivée'] = df['Arrivée'].dt.strftime('%d%H%M')
    return df

def completer_pavillon(df1, df2):
    df1 = df1.merge(df2[['Lloyd', 'Pavillon']], on='Lloyd', how='left')
    df1['Pavillon'] = df1['Pavillon'].fillna("  ")
    return df1

def convertir_pavillons(df, colonne):
    # version courte
    df[colonne] = df[colonne].str.lower().map({
        'france': 'FR', 'germany': 'DE', 'belgium': 'BE', 'netherlands': 'NL',
        'spain': 'ES', 'italy': 'IT', 'united kingdom': 'GB'
    }).fillna('  ')

def transcrire_type(type_donnees):
    correspondance = {
        'Cargo': 'TM', 'Vraquier': 'TMB', 'Porte Conteneur ': 'TMC', 'Gazier': 'TMOT',
        'Chimiquier': 'TMOS', 'Pétrolier': 'TMO', 'Paquebot': 'TMP',
    }
    return correspondance.get(type_donnees)

def supprimer_doublons(df):
    return df.drop_duplicates()

def modifier_contenu(df, nom_colone, value):
    if nom_colone not in df.columns:
        return df
    df[nom_colone] = value
    return df


#bloc 3 FONCTION SPECIFIQUE HAVRE

def afficher_donnees_dans_plage(df, date_debut, date_fin):
    date_debut = datetime.strptime(str(date_debut), "%Y-%m-%dT%H:%M")
    date_fin = datetime.strptime(str(date_fin), "%Y-%m-%dT%H:%M")
    date_debut = pd.to_datetime(date_debut.strftime("%d/%m/%Y %H:%M"), dayfirst=True)
    date_fin = pd.to_datetime(date_fin.strftime("%d/%m/%Y %H:%M"), dayfirst=True)
    df['Arrivée'] = pd.to_datetime(df['Arrivée'], dayfirst=True)
    return df[(df['Arrivée'] >= date_debut) & (df['Arrivée'] <= date_fin)]

def parenthese(df):
    return df[~df['NOM'].str.contains(r'\(.*\)', regex=True)]

def replace_with_first_letters(df, column_name):
    df[column_name] = df[column_name].apply(lambda x: x[:3] if isinstance(x, str) else x)
    return df

def convert_country_codes(df, name):
    def alpha3_to_alpha2(alpha3):
        try:
            return pycountry.countries.get(alpha_3=alpha3).alpha_2
        except:
            return '  '
    df[name] = df[name].apply(alpha3_to_alpha2)
    return df

def transcrire_type_donnees(type_donnees):
    transcriptions = {
        'TCR': 'TMO', 'TPD': 'TMO', 'UCC': 'TMC', 'PRR': 'TMF',
        'TCH': 'TMOS', 'DHD': 'TMD', 'MVE': 'TME', 'MPR': 'TMP',
    }
    return transcriptions.get(type_donnees)


#bloc 4 format excel

def format_excel(data1, data2, port_name, nom):
    wb = Workbook()
    ws = wb.active

    # Titre principal
    ws.merge_cells('A1:I1')
    ws['A1'] = 'PREVISIONS DE MOUVEMENTS A 72 HEURES*'
    ws['A1'].font = Font(size=24)
    ws['A1'].alignment = Alignment(horizontal='left')

    # Avertissement en rouge
    ws.merge_cells('A2:I2')
    ws['A2'] = '* Ne pas mentionner les navires à passagers effectuant des liaison régulières'
    ws['A2'].font = Font(color="FF0000", italic=True)
    ws['A2'].alignment = Alignment(horizontal='left')

    # Informations générales
    ws['A5'] = 'ATTENDU:'
    ws['B5'] = f'{port_name} 72H'
    ws['C5'] = len(data1.index)

    # Colonnes pour df1
    for c_idx, col_name in enumerate(data1.columns, start=2):
        cell = ws.cell(row=7, column=c_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for r_idx, row in enumerate(data1.values, start=8):
        for c_idx, value in enumerate(row, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    # Section DEPART
    last_row = len(data1.index) + 15
    ws.cell(row=last_row, column=1, value='A QUAI:')
    ws.cell(row=last_row, column=2, value=f'{port_name} 72H')
    ws.cell(row=last_row, column=3, value=len(data2.index))

    for c_idx, col_name in enumerate(data2.columns, start=2):
        cell = ws.cell(row=last_row + 2, column=c_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for r_idx, row in enumerate(data2.values, start=last_row + 3):
        for c_idx, value in enumerate(row, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    # Ajustement des colonnes
    for column_cells in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Coloration jaune pour valeurs manquantes
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == '  ':
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Bordures
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Note en bas
    last_row = len(data1.index) + len(data2.index) + 30
    ws.merge_cells(f'A{last_row}:I{last_row}')
    ws.cell(row=last_row, column=1, value='* *Préciser si le navire est prévu au mouillage en zone d\'attente')
    ws.cell(row=last_row, column=1).font = Font(color="FF0000", italic=True)

    wb.save(os.path.join("generated", nom))

def get_remote_driver():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')

    remote_url = os.getenv("SELENIUM_REMOTE_URL", "http://selenium:4444/wd/hub")

    return webdriver.Remote(
        command_executor=remote_url,
        options=options
    )


def create_df2(lien_page, xpath_colonne_bateau, id_table):
    driver = get_remote_driver()
    driver.get(lien_page)
    try:
        # On lit directement le tableau par son ID depuis le HTML
        df = pd.read_html(StringIO(driver.page_source), attrs={'id': str(id_table)}, header=0)[0]
        driver.quit()
        return df
    except Exception as e:
        driver.quit()
        raise Exception(f"Erreur lors de la lecture du tableau HTML : {str(e)}")


# app/services/utils.py

import pycountry

def convert_country_to_iso2(country_name):
    try:
        return pycountry.countries.lookup(country_name).alpha_2
    except Exception:
        return ''
